from flask import Flask, render_template, request, send_file
import os
import xml.etree.ElementTree as ET
import pandas as pd
import xlsxwriter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])

def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            print('Aucun fichier envoyé.')  # Impression de débogage
            return 'Aucun fichier envoyé.'

        file = request.files['file']
        if file.filename == '':
            print('Aucun fichier sélectionné.')  # Impression de débogage
            return 'Aucun fichier sélectionné.'

        if file and file.filename.endswith('.xml'):
            xml_file_path = os.path.join('uploads', file.filename)
            file.save(xml_file_path)
            print(f'Fichier sauvegardé à {xml_file_path}') 
            excel_file_path = xml_to_excel(xml_file_path)
            return send_file(excel_file_path, as_attachment=True)

    return render_template('index.html')

def xml_to_excel(xml_file_path):
    print(f"Tentative de traitement du fichier XML: {xml_file_path}")
    if not os.path.exists(xml_file_path):
        print("Le fichier XML n'existe pas.")
        return None

    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    namespaces = {'ns': 'http://www.plmxml.org/Schemas/PLMXMLSchema'}
    
    # Extraire toutes les données
    data = []
    for element in root.findall('.//ns:*', namespaces):
        # Obtenir le nom de l'élément
        element_name = element.tag.split('}')[-1]
        
        # Ignorer l'élément si c'est 'Form'
        if element_name == 'Form' :
            continue
        
        if element_name == 'AccessIntent' :
            continue
        if element_name == 'VariantCondition' :
            continue
        if element_name == 'InstanceGraph' :
            continue
        if element_name == 'AccessIntent' :
            continue
        if element_name == 'Occurrence' :
            continue
        if element_name == 'ProductDef' :
            continue
        if element_name == 'ProductRevisionView' :
            continue
        if element_name == 'ProductView' :
            continue
        if element_name == 'Site' :
            continue
        if element_name == 'Transform' :
            continue
        if element_name == 'VariantCondition' :
            continue
        
        element_data = {'Element': element_name}
        element_data.update(element.attrib)
        
        if element.text:
            element_data['Text'] = element.text.strip()
        
        data.append(element_data)
    
    # Créer un DataFrame
    df = pd.DataFrame(data)
    
    # Créer un fichier Excel et une feuille
    wb = Workbook()
    ws = wb.active

    # Ajouter les données du DataFrame au fichier Excel
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
        
# Obtenir la plage de données pour le tableau (inclure les en-têtes)
    min_col = 1
    max_col = df.shape[1]
    min_row = 1
    max_row = len(df) + 1  # +1 pour inclure les en-têtes

# Créer un objet Table
    tab_range = f"A{min_row}:{get_column_letter(max_col)}{max_row}"
    tab = Table(displayName="Table1", ref=tab_range)

    # Créer un style de tableau (vous pouvez choisir n'importe quel style prédéfini)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    # Ajouter la table à la feuille de calcul
    ws.add_table(tab)

    
    # Sauvegarder le DataFrame dans un fichier Excel
    excel_file_path = xml_file_path.replace('.xml', '.xlsx')
    df.to_excel(excel_file_path, index=False)
    return excel_file_path

if __name__ == '__main__':
    app.run(debug=True)
